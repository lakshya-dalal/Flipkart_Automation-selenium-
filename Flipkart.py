from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
#next two lines to get incognito
chrome_option = webdriver.ChromeOptions()
chrome_option.add_argument('--incognito')
chrome_option.add_argument("headless")
driver=webdriver.Chrome(executable_path="C:\\chromedriver.exe", options=chrome_option)
driver.implicitly_wait(5)
driver.maximize_window()#to open in maximise
import openpyxl
book =openpyxl.load_workbook("C:\\Users\\Lakshya\\Desktop\\nn.xlsx")
sheet =book.active
for i in range(645,sheet.max_row+1):  # to get rows sheet.max_row+1
    for j in range(1,2):#to get columns sheet.max_column+1
        id=str(sheet.cell(row=i, column=j).value)
        pas=str(sheet.cell(row=i, column=j+1).value)
        driver.get("https://www.flipkart.com/account/orders?link=home_orders")
        wait = WebDriverWait(driver, 15)
        wait.until(expected_conditions.presence_of_element_located((By.CSS_SELECTOR, "input[class='_2IX_2- VJZDxU']")))
        driver.find_element_by_css_selector("input[class='_2IX_2- VJZDxU']").send_keys(id)#username
        driver.find_element_by_css_selector("input[class='_2IX_2- _3mctLh VJZDxU']").send_keys(pas)
        driver.find_element_by_css_selector("button[class='_2KpZ6l _2HKlqd _3AWRsL']").click()
        try:
            if "incorrect" in driver.find_element_by_xpath("//span[@class='_2YULOR']/span").text:
                sheet.cell(row=i, column=j+2).value="error in login"
        except:
            #print no. of orders
            names=driver.find_elements_by_xpath("//span[@class='row _1kkfO3 BqOr_g']")
            if(len(names)==0):
                sheet.cell(row=i, column=j+2).value="no orders"
            else:
                delivery=driver.find_elements_by_xpath("//span[@class='AO0UbU']")
                count=0
                for name,singledel in zip(names,delivery):
                    if(name.text.find("GB") != -1 and singledel.text.find("eliver")!=-1):
                        count=count+1
                    else :
                        continue
                for name,singledel in zip(names,delivery):
                    if(name.text.find("GB") != -1 and singledel.text.find("eliver")!=-1):
                        sheet.cell(row=i, column=j+5).value=singledel.text[13:]
                        name.click()
                        wait.until(expected_conditions.presence_of_element_located((By.XPATH, "//div[@class='_2NKhZn']/p")))
                        orderid=driver.find_element_by_xpath("//div[@class='_2NKhZn']/p")
                        sheet.cell(row=i, column=j+2).value=orderid.text
                        company=driver.find_element_by_xpath("//div[@class='_10YZ6u']")
                        sheet.cell(row=i, column=j+3).value=company.text[0:4]
                        break
                if(count==0):
                    sheet.cell(row=i, column=j+2).value="cancelled"
                else:
                    sheet.cell(row=i, column=j+4).value=count
        book.save("C:\\Users\\Lakshya\\Desktop\\nn.xlsx")
        driver.delete_all_cookies()
driver.close()
